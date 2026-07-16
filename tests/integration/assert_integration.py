#!/usr/bin/env python3
"""Assert the mock run produced the expected resolution, funnel, and calls."""
from __future__ import annotations

import argparse
import csv
import os
import sys

from cyvcf2 import VCF

FAILS = []


def check(cond, msg):
    print(("PASS " if cond else "FAIL ") + msg)
    if not cond:
        FAILS.append(msg)
    return cond


def rows(path):
    with open(path) as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--work", required=True)
    a = ap.parse_args(argv)
    W = a.work

    # --- resolution ---
    res = {r["kid"]: r for r in rows(os.path.join(W, "trio_resolution.tsv"))}
    check(res.get("CH_A", {}).get("status", "").startswith("resolved"), "CH_A resolved")
    check(res.get("CH_B", {}).get("status", "").startswith("resolved"), "CH_B resolved (family VCF, extra sib)")
    check(res.get("CH_C", {}).get("status") == "unresolved", "CH_C unresolved")
    check("MO_C" in res.get("CH_C", {}).get("missing_members", ""), "CH_C reports missing mom MO_C")
    manifest = rows(os.path.join(W, "trios.resolved.tsv"))
    check(len(manifest) == 2, f"2 trios in resolved manifest (got {len(manifest)})")

    # --- Step 0: CH_B inferred male + contamination gate ---
    qc = {r["trio_id"]: r for r in rows(os.path.join(W, "qc_report.tsv"))}
    check(qc.get("CH_B", {}).get("inferred_sex") == "1", "CH_B inferred male (chrX)")
    # no selfSM configured -> VCF-only CHARR fallback; mock hom-alt AD is 0 ref -> ~0, unflagged
    check(qc.get("CH_A", {}).get("contam_source") == "charr", "contamination falls back to CHARR")
    check(qc.get("CH_A", {}).get("contam_flag") == "0", "CH_A not flagged contaminated (clean)")
    # CH_B's father carries 6 ref reads at a hom-alt site -> CHARR 0.15 > 0.02 -> trio flagged
    check(qc.get("CH_B", {}).get("contam_flag") == "1", "CH_B flagged contaminated (father CHARR)")
    dadc = float(qc.get("CH_B", {}).get("dad_contam") or 0)
    check(0.1 < dadc < 0.2, f"CH_B dad_contam ~0.15 (got {dadc})")
    check((qc.get("CH_B", {}).get("kid_contam") or "0") in ("", "0"), "CH_B proband CHARR clean")
    # Mendelian-error gate (sample-swap proxy): the de novo in CH_A is a Mendelian violation
    check(int(qc.get("CH_A", {}).get("mie_errors") or 0) >= 1, "CH_A Mendelian error detected (de novo)")
    check(qc.get("CH_A", {}).get("mie_flag") == "1", "CH_A MIE flag raised")

    # --- Step 3: plausible sites keep/drop ---
    plaus = {}
    for v in VCF(os.path.join(W, "plausible.sites.vcf.gz")):
        plaus[(v.CHROM, v.POS)] = v.INFO.get("hprv_keep_reason")
    check(("chr2", 8000) in plaus and plaus[("chr2", 8000)] == "clinvar_plp",
          "ClinVar P/LP (LOW impact) kept via clinvar_plp — on UNSTARRED CLIN_SIG from the VEP "
          "cache (there is no CLNREVSTAT, so the old >=2-star gate cannot and does not apply)")
    check(("chr1", 12000) not in plaus, "BA1-common variant dropped at Step 3")
    check(("chr1", 17000) not in plaus, "non-PASS variant dropped before Step 3")
    check(("chr1", 5000) in plaus, "de novo site retained as plausible")

    # --- VEP-only contract: CADD is the ONLY functional predictor, hence the ONLY way any
    # variant below MODERATE impact can survive. If this regresses the screen silently goes
    # coding-only and every intronic/synonymous candidate vanishes. ---
    check(plaus.get(("chr1", 18000)) == "cadd",
          "deep-intronic MODIFIER kept via CADD (the sole non-coding keep-path)")
    check(("chr1", 18500) not in plaus,
          "intronic MODIFIER with sub-threshold CADD dropped (the CADD gate really gates)")

    # --- The gating-dead predictors must not reappear. REVEL/AlphaMissense/MPC are missense-only
    # scores, and every missense is IMPACT=MODERATE, which selection.py keeps at an earlier
    # branch — so these keep-reasons were unreachable even when the code still had them. This is
    # the audit's falsifiable prediction, enforced: they must be exactly 0, cohort-wide. ---
    for dead in ("revel", "alphamissense", "mpc", "spliceai", "loftee_hc"):
        check(dead not in set(plaus.values()), f"no site kept via '{dead}' (removed / unreachable)")

    # --- Step 5: inheritance calls ---
    calls = rows(os.path.join(W, "candidates.calls.tsv"))

    def has(trio, mode, chrom=None, pos=None, gene=None):
        for r in calls:
            if r["trio_id"] == trio and r["mode"] == mode \
               and (chrom is None or r["chrom"] == chrom) \
               and (pos is None or r["pos"] == str(pos)) \
               and (gene is None or r["symbol"] == gene):
                return True
        return False

    check(has("CH_A", "denovo", "chr1", 5000, "GENE1"), "CH_A de novo GENE1")
    check(has("CH_A", "hom_recessive", "chr1", 8000, "GENE2"), "CH_A hom recessive GENE2")

    # --- MULTIALLELIC trans comp-het (child 1/2). `norm -m-` splits it and, without
    # --keep-sum AD, strips the other ALT's reads -> ref_ad~0 -> allele_balance ~1.0 -> the het
    # band rejects BOTH legs and the pair vanishes silently. This is a MISSED DIAGNOSIS channel:
    # it hits exactly the loci where comp-hets concentrate, and no counter records the loss. ---
    ch2 = [r for r in calls if r["trio_id"] == "CH_A" and r["symbol"] == "GENECH2"]
    check(len([r for r in ch2 if r["mode"] == "compound_het"]) == 2,
          "multiallelic (1/2) child yields a compound_het PAIR in GENECH2 — the --keep-sum AD fix")
    check(len({r["pair_id"] for r in ch2 if r["mode"] == "compound_het"}) == 1,
          "both GENECH2 legs share one pair_id (a genuine trans pair, not two singletons)")
    # ...and the child's allele balance must be the corrected ~0.5, not the 1.0 the bug produced
    for r in ch2:
        ab = float(r["child_ab"]) if r["child_ab"] else 0.0
        check(0.25 <= ab <= 0.75,
              f"GENECH2 leg chr1:{r['pos']} child_ab={ab:.3f} inside the het band (bug gave 1.000)")
    # scoped to GENE3: CH_A now also has a legitimate multiallelic comp-het in GENECH2 (below),
    # so an unscoped count of every CH_A compound_het row would be 4, not 2.
    ch = [r for r in calls if r["trio_id"] == "CH_A" and r["mode"] == "compound_het"
          and r["symbol"] == "GENE3"]
    check(len(ch) == 2, f"CH_A compound het pair in GENE3 (got {len(ch)})")
    # comp-het requires TRANS: two cis (both maternal) hets in GENEC must NOT be a compound_het
    genec = [r for r in calls if r["trio_id"] == "CH_A" and r["symbol"] == "GENEC"]
    check(genec and all(r["mode"] != "compound_het" for r in genec),
          "cis (same-parent) GENEC pair NOT called compound_het")
    check(any(r["mode"] == "dominant" for r in genec), "cis GENEC variants emitted as dominant instead")
    check(has("CH_B", "denovo", "chr1", 5100, "GENE1"), "CH_B de novo GENE1 (secondary)")
    check(has("CH_B", "x_linked_recessive", "chrX", 2781600, "GENEX"), "CH_B X-linked recessive GENEX")
    # X-linked recessive must fire even with an AFFECTED (hom-alt) father (father's chrX not
    # transmitted to a son) — the previously-required father-hom-ref would have wrongly dropped it
    check(has("CH_B", "x_linked_recessive", "chrX", 2782000, "GENEXAF"),
          "X-linked recessive called with an affected (hom-alt) father")
    # autosomal hom-recessive with a HOM-ALT parent (carrier rule accepts HET or HOM_ALT parents)
    check(has("CH_A", "hom_recessive", "chr1", 8500, "GENE2H"), "hom recessive called with a HOM-ALT parent")
    check(not has("CH_A", "denovo", "chr1", 15000), "low-GQ pseudo-de-novo NOT called (QC gate)")
    # dominant model: rare functional inherited het, recurrent across individuals
    check(has("CH_A", "dominant", "chr2", 10000, "GENED"), "CH_A dominant inherited het GENED")
    check(has("CH_B", "dominant", "chr2", 10000, "GENED"), "CH_B dominant inherited het GENED")
    # Rarity oracle = grpmax PROXY, not VEP's MAX_AF. GENEFND is at AF 0.002 in gnomAD 'mid' — a
    # bottlenecked group gnomAD's own grpmax excludes — and absent from every grpmax-eligible
    # group. MAX_AF therefore reports 0.002, 20x over dominant_max=1e-4, and reading it would
    # silently DROP this call. frequency() must see None and the call must survive.
    check(has("CH_A", "dominant", "chr2", 17000, "GENEFND"),
          "founder-population-only allele (MAX_AF=0.002 in 'mid') still called dominant — "
          "grpmax proxy correctly ignores bottlenecked groups")
    fnd = [r for r in calls if r["symbol"] == "GENEFND"]
    check(fnd and all(not r["grpmax_af"] for r in fnd),
          "GENEFND reports an EMPTY grpmax_af (no eligible group carries it)")
    check(fnd and all(r["max_af"] for r in fnd),
          "GENEFND still reports max_af for the reviewer to see why it looked common")
    # CADD-only keep survives into an actual call, not just Step 3
    check(has("CH_A", "dominant", "chr1", 18000, "GENEIN"),
          "deep-intronic CADD-kept variant becomes a dominant call")

    # --- Step 6: recurrence-based gene consolidation ---
    genes = {r["gene"]: r for r in rows(os.path.join(W, "genes.ranked.tsv"))}
    check(genes.get("GENED", {}).get("n_dominant") == "2", "GENED has 2 dominant carriers")
    check(genes.get("GENED", {}).get("recurrent") == "1", "GENED flagged recurrent")
    # same- vs distinct-variant recurrence: GENED shares one variant (founder/artifact-suspect);
    # GENEDD has two distinct variants across trios (the stronger gene signal)
    check(genes.get("GENED", {}).get("recurrence_kind") == "same_variant", "GENED = same-variant recurrence")
    check(genes.get("GENEDD", {}).get("recurrence_kind") == "distinct_variant", "GENEDD = distinct-variant recurrence")
    check(genes.get("GENEDD", {}).get("n_dominant") == "2", "GENEDD has 2 dominant carriers")
    check(genes.get("GENE1", {}).get("n_denovo") == "2", "GENE1 has 2 de novo carriers (secondary)")
    # calibrated recurrence null: a rare variant recurring in 2 individuals is significant
    check(float(genes.get("GENED", {}).get("p_recurrence") or 1) < 1e-4,
          "GENED has a small calibrated recurrence p-value")
    check(genes.get("GENED", {}).get("recurrence_exome_wide_sig") == "1",
          "GENED recurrence is exome-wide significant")
    # GENE1 (de novo only) must NOT get an inherited recurrence p-value
    check(not genes.get("GENE1", {}).get("p_recurrence"),
          "GENE1 (de novo only) has no inherited recurrence p-value")

    # --- chrM is OUT OF SCOPE and must never reach any output. The mock carries a near-fixed
    # rCRS haplogroup variant (m.8860A>G, whole trio hom-alt) in BOTH trios. Un-excluded it fires
    # hom_recessive everywhere and — with no gnomAD mito AF to fail the rarity gate — floors q in
    # Step 6 and lands in the recurrent, exome-wide-significant tier above real nuclear genes. ---
    check(not any(r["chrom"] in ("chrM", "chrMT", "M", "MT") for r in calls),
          "no chrM call in candidates.calls.tsv (chrM excluded at Step 1)")
    check("MT-ATP6" not in genes,
          "MT-ATP6 absent from genes.ranked.tsv — a haplogroup variant never reaches the "
          "recurrent tier")
    for v in VCF(os.path.join(W, "cohort.sites.vcf.gz")):
        if v.CHROM in ("chrM", "chrMT", "M", "MT"):
            check(False, f"chrM leaked into the cohort union at {v.CHROM}:{v.POS}")
            break
    else:
        check(True, "cohort.sites.vcf.gz contains no chrM records")

    # --- GATK's de novo tags must SURVIVE Step 4. A blanket `annotate -x INFO` stripped them,
    # which silently made Step 5's has_hiconf permanently False and filters.denovo.use_hiconf_tag
    # a no-op — invisible, because de novo is still called when the tag is simply absent. ---
    trio_vcfs = {r["trio_id"]: r["candidates_vcf"] for r in rows(os.path.join(W, "trios.candidates.tsv"))}
    if check("CH_A" in trio_vcfs, "CH_A candidate VCF in the manifest"):
        hdr = VCF(trio_vcfs["CH_A"]).raw_header
        check("ID=hiConfDeNovo" in hdr,
              "hiConfDeNovo header SURVIVES Step 4's INFO strip (else the de novo gate is dead code)")
        check(any(v.INFO.get("hiConfDeNovo") for v in VCF(trio_vcfs["CH_A"])),
              "at least one candidate record still carries a hiConfDeNovo value")
        check(any(str(k).startswith("vep_") for v in VCF(trio_vcfs["CH_A"]) for k, _ in v.INFO),
              "vep_* annotations still transfer into the per-trio VCF")

    # --- audit exists ---
    check(os.path.exists(os.path.join(W, "audit", "summary.md")), "audit/summary.md written")
    counts = rows(os.path.join(W, "audit", "counts.tsv"))
    check(any(r["step"] == "resolve" and r["metric"] == "trios_resolved" and r["value"] == "2"
              for r in counts), "audit records 2 resolved trios")

    # --- Step 7: xlsx summary ---
    xlsx = os.path.join(W, "hprv_summary.xlsx")
    if check(os.path.exists(xlsx), "xlsx summary written"):
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, read_only=True)
        for sh in ("About", "Gene consolidation", "Candidate calls"):
            check(sh in wb.sheetnames, f"xlsx has '{sh}' sheet")

    # --- Step 8: igv.js variant-review export ---
    vpath = os.path.join(W, "igv", "variants.tsv")
    if check(os.path.exists(vpath), "igv variants.tsv written"):
        vh, vrows = None, []
        with open(vpath) as fh:
            rr = list(csv.reader(fh, delimiter="\t"))
        vh, vrows = rr[0], rr[1:]
        for col in ("chrom", "pos", "ref", "alt", "inheritance", "child_file", "child_gt"):
            check(col in vh, f"variants.tsv has '{col}' column")
        check(len(vrows) == len(calls), f"variants.tsv rows == candidate calls ({len(calls)})")
        fi = vh.index("child_file")
        check(any(r[fi] for r in vrows), "at least one child_file (mini-CRAM) populated")
    check(os.path.exists(os.path.join(W, "igv", "crams", "CH_A", "CH_A.cram")),
          "CH_A mini-CRAM extracted")
    check(os.path.exists(os.path.join(W, "igv", "trios.tsv")), "igv trios.tsv written")
    check(os.path.exists(os.path.join(W, "igv", "curation.json")), "igv curation.json written")

    if FAILS:
        sys.stderr.write(f"\n{len(FAILS)} assertion(s) FAILED\n")
        return 1
    print("\nALL INTEGRATION ASSERTIONS PASSED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
