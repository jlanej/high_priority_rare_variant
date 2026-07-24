"""Build a consolidated .xlsx summary of a run — a documented supplemental table.

Sheets: About (legend + run summary + thresholds), Gene consolidation (the headline
recurrence-ranked table), Candidate calls (per-variant with inheritance mode +
genotypes + annotations), Trio resolution, QC, and the raw Audit counts. The values
are analysis RESULTS (not a calculation model), so they are written as data — there
are no spreadsheet formulas to recalculate.
"""

from __future__ import annotations

import csv
import gzip
import math
import os
import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hprv.config import get

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
H2_FONT = Font(name=FONT, bold=True, size=11)
BASE_FONT = Font(name=FONT)


def _coerce(tok):
    """Coerce a TSV token to int/float so numeric columns SORT numerically under the sheet's
    auto-filter. openpyxl stores a Python str as a TEXT cell, which Excel sorts lexicographically
    ('0.0013' before '2.5e-06', '1.2e-08' after '0.05') — so the rarest / most-significant rows do
    NOT sort to the top on a user re-sort, and every numeric cell shows the "number stored as text"
    warning. Non-numeric tokens (gene / chrom / GT like '0/1' / IDs / flags / '' / '.') fail int()+
    float() and pass through unchanged; non-finite floats (nan/inf) stay text defensively."""
    if not tok:
        return tok
    try:
        return int(tok)
    except (ValueError, TypeError):
        pass
    try:
        f = float(tok)
    except (ValueError, TypeError):
        return tok
    return f if math.isfinite(f) else tok


def _vep_header_line(vcf_path):
    """Return the first ``##VEP=`` header line of a (bgzipped) annotated VCF, or ''."""
    if not vcf_path or not os.path.exists(vcf_path):
        return ""
    try:
        with gzip.open(vcf_path, "rt") as fh:
            for ln in fh:
                if not ln.startswith("#"):
                    break
                if ln.startswith("##VEP="):
                    return ln.rstrip("\n")
    except OSError:
        return ""
    return ""


def _gnomad_label(vep_line, default="gnomAD v4.1"):
    """Derive a 'gnomAD vX' label from the ``##VEP=`` line rather than hardcoding it. The VEP cache
    version (r113+) determines the gnomAD release, so under the enforced self-run path the default is
    correct-by-construction; this only avoids over-claiming on a non-v4.1 ingest VCF."""
    import re
    m = re.search(r'gnomAD[eg]?[=_ ]v?([0-9]+\.[0-9]+)', vep_line or "")
    return f"gnomAD v{m.group(1)}" if m else default


def _read_tsv(path):
    if not path or not os.path.exists(path):
        return None, []
    with open(path) as fh:
        rows = list(csv.reader(fh, delimiter="\t"))
    if not rows:
        return None, []
    return rows[0], rows[1:]


def _style_data_sheet(ws, header, rows, freeze=True):
    ws.append(header)
    for r in rows:
        ws.append([_coerce(x) for x in r])   # typed numeric cells so the auto-filter sorts correctly
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    # column widths (capped), base font on data
    for c in range(1, len(header) + 1):
        col = get_column_letter(c)
        maxlen = len(str(header[c - 1]))
        for r in rows[:500]:
            if c - 1 < len(r):
                maxlen = max(maxlen, len(str(r[c - 1])))
        ws.column_dimensions[col].width = min(max(maxlen + 2, 8), 48)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = BASE_FONT
    if freeze:
        ws.freeze_panes = "A2"
    if header and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"


def _tally_modes(header, rows):
    if not header or "mode" not in header:
        return {}
    i = header.index("mode")
    out = {}
    for r in rows:
        if i < len(r):
            out[r[i]] = out.get(r[i], 0) + 1
    return out


def build(work_dir, out_xlsx, cfg, run_label=""):
    p = lambda *a: os.path.join(work_dir, *a)  # noqa: E731
    calls_h, calls = _read_tsv(p("candidates.calls.tsv"))
    genes_h, genes = _read_tsv(p("genes.ranked.tsv"))
    res_h, res = _read_tsv(p("trio_resolution.tsv"))
    qc_h, qc = _read_tsv(p("qc_report.tsv"))
    audit_h, audit = _read_tsv(p("audit", "counts.tsv"))

    # Per-run annotation provenance (P9): derive from artifacts/config rather than hardcoding, so the
    # frequency oracle (golden rule 2) is RECORDED, not asserted. The ##VEP= line lands at the same
    # path on both the self-run and ingest paths; the gnomAD label is derived from it (correct-by-
    # construction on the self-run cache; no over-claim on a non-v4.1 ingest VCF).
    vep_line = _vep_header_line(p("cohort.sites.annotated.vcf.gz"))
    gnomad_label = _gnomad_label(vep_line)

    wb = Workbook()

    # ---- About / legend ----
    ws = wb.active
    ws.title = "About"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90

    def line(a="", b="", bold=False, title=False, h2=False):
        r = ws.max_row + 1 if ws.max_row > 1 or ws["A1"].value else 1
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        f = TITLE_FONT if title else (H2_FONT if h2 else (Font(name=FONT, bold=bold)))
        ws.cell(row=r, column=1).font = f
        ws.cell(row=r, column=2).font = BASE_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    line("high_priority_rare_variant", "Consolidated analysis summary", title=True)
    if run_label:
        line("Run", run_label)
    line("", "")
    line("Purpose", "Screens GMKF Kids First per-trio VCFs (GRCh38) for high-priority INHERITED "
                    "rare variants and consolidates genes where rare functional variants recur "
                    "across individuals. De novo filtering/review and mtDNA heteroplasmy are handled "
                    "by separate dedicated pipelines (de novo is a secondary cross-reference here).")
    line("", "")

    line("Sheets", "", h2=True)
    for nm, desc in [
        ("Gene consolidation", "Genes ranked by recurrence across individuals (dominant het / "
                               "biallelic / X-linked), weighted by constraint. The headline result."),
        ("Candidate calls", "One row per candidate per trio: inheritance mode, genotypes, and "
                            "annotations (gnomAD grpmax AF, CADD, ClinVar)."),
        ("Trio resolution", "Which VCF each kid/dad/mom trio resolved to; unresolved trios + why."),
        ("QC", "Per-trio Mendelian-error rate, chrX-inferred sex, and contamination "
               "(verifyBamID FREEMIX or VCF-only CHARR) — the garbage-in guard."),
        ("Audit counts", "Per-step input/output counts and funnel tallies (what went where, and why)."),
    ]:
        line(nm, desc)
    line("", "")

    # run summary
    line("Run summary", "", h2=True)
    modes = _tally_modes(calls_h, calls)
    n_resolved = sum(1 for r in res if res_h and "status" in res_h
                     and r[res_h.index("status")].startswith("resolved")) if res_h else 0
    n_recurrent = sum(1 for r in genes if genes_h and "recurrent" in genes_h
                      and r[genes_h.index("recurrent")] == "1") if genes_h else 0
    line("Trios resolved", str(n_resolved))
    line("Candidate calls", str(len(calls)))
    line("Calls by mode", ", ".join(f"{k}={v}" for k, v in sorted(modes.items())) or "(none)")
    line("Genes nominated", str(len(genes)))
    line("Recurrent genes", f"{n_recurrent} (>= {get(cfg, 'burden.min_carriers', 2)} distinct individuals)")
    # significant in ANY inherited family (dominant / biallelic / X-linked), matching genes.ranked.tsv
    _sig_cols = [c for c in ("recurrence_exome_wide_sig", "recurrence_biallelic_exome_wide_sig",
                             "recurrence_xlinked_exome_wide_sig") if genes_h and c in genes_h]
    n_recsig = sum(1 for r in genes
                   if any(r[genes_h.index(c)] == "1" for c in _sig_cols)) if genes_h else 0
    line("Recurrence-significant genes",
         f"{n_recsig} (exome-wide p < {get(cfg, 'burden.exome_wide_p', 2.5e-6):g} on the calibrated "
         "recurrence null; any inherited model — dominant / biallelic / X-linked)")
    line("", "")

    # key thresholds
    line("Key thresholds (configurable defaults)", "", h2=True)
    line("Rarity field", f"{gnomad_label} grpmax-proxy AF = max AF over the grpmax-eligible ancestry "
                         "groups (AFR/AMR/EAS/NFE/SAS), from the VEP cache. A POINT ESTIMATE: "
                         "faf95 (the 95% CI lower bound) needs AC/AN, which the cache does not "
                         "carry, so these gates run slightly stringent on low-count alleles.")
    line("Dominant / de novo rarity", f"grpmax AF < {get(cfg, 'filters.rarity.dominant_max', 1e-4)}")
    line("Recessive rarity", f"grpmax AF < {get(cfg, 'filters.rarity.recessive_max', 1e-2)}")
    line("Benign (never rescued)", f"grpmax AF >= {get(cfg, 'filters.rarity.benign_ba1', 0.05)} (ClinGen BA1)")
    line("Genotype QC", f"GQ >= {get(cfg, 'filters.genotype_qc.min_gq', 20)}, "
                        f"DP >= {get(cfg, 'filters.genotype_qc.min_dp', 10)}, "
                        f"het AB {get(cfg, 'filters.genotype_qc.het_ab_min', 0.25)}-"
                        f"{get(cfg, 'filters.genotype_qc.het_ab_max', 0.75)}")
    line("Recurrence", f"gene flagged recurrent at >= {get(cfg, 'burden.min_carriers', 2)} distinct individuals")
    line("", "")
    line("Notes", "Thresholds are configurable; a gene-specific ClinGen VCEP value overrides a "
                  "generic cutoff. See the repo docs/ for calibrated methods and citations.")
    line("", "")

    # ---- provenance (P9): make the run bindable from the workbook itself ----
    line("Provenance", "", h2=True)
    line("Frequency oracle", gnomad_label + " (from the VEP cache; the sole population-frequency source)")
    line("VEP", f"release {get(cfg, 'resources.vep.version', '115')}; cache "
                f"{os.path.basename(str(get(cfg, 'resources.vep.cache_dir', '')).rstrip('/')) or '(unset)'}")
    cadd_snv = os.path.basename(str(get(cfg, 'resources.vep.cadd_snv', '')))
    cadd_indel = os.path.basename(str(get(cfg, 'resources.vep.cadd_indel', '')))
    line("CADD files", ", ".join(x for x in (cadd_snv, cadd_indel) if x) or "(not used)")
    git_sha = ""
    try:
        git_sha = subprocess.run(["git", "-C", os.path.dirname(os.path.abspath(__file__)),
                                  "rev-parse", "--short", "HEAD"],
                                 capture_output=True, text=True, check=False, timeout=5).stdout.strip()
    except (OSError, ValueError):
        git_sha = ""
    line("Pipeline commit", git_sha or "(unavailable — not a git checkout / baked image)")
    if run_label:
        line("Run label", run_label)
    if vep_line:
        # the verbatim VEP header line — the definitive, self-describing provenance record
        line("VEP header", vep_line[:400])
    ws.sheet_view.showGridLines = False

    # ---- data sheets ----
    for title, header, rows in [
        ("Gene consolidation", genes_h, genes),
        ("Candidate calls", calls_h, calls),
        ("Trio resolution", res_h, res),
        ("QC", qc_h, qc),
        ("Audit counts", audit_h, audit),
    ]:
        if not header:
            continue
        _style_data_sheet(wb.create_sheet(title), header, rows)

    wb.save(out_xlsx)
    return out_xlsx
